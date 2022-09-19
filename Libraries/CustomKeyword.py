import os
import json
import ast
import sys
import pyttsx3
import openpyxl
from robot.api.deco import library, keyword
import pandas as pd
import speech_recognition as sr
from pydub import AudioSegment
import pyautogui


@library
class CustomKeyword:

    # Write data to excel
    @keyword
    def write_data_excel(self, sheet_name, scenario_name, header_name, value_to_write):
        book = openpyxl.load_workbook("TestData/CsvFiles/Test_Data.xlsx")
        sheet = book[sheet_name]
        print(os.path)
        for i in range(1, sheet.max_row + 1):
            row_scenario = sheet.cell(row=i, column=1)
            if row_scenario.value == scenario_name:
                for j in range(1, sheet.max_column + 1):
                    column_header = sheet.cell(row=1, column=j)
                    if column_header.value == header_name:
                        sheet.cell(row=i, column=j).value = value_to_write
                        book.save("TestData/CsvFiles/Test_Data.xlsx")

    # Read the data from Excel
    @keyword
    def read_data_excel(self, sheet_name, scenario_name, header_name):
        book = openpyxl.load_workbook("TestData/CsvFiles/Test_Data.xlsx")
        sheet = book[sheet_name]
        for i in range(1, sheet.max_row + 1):
            row_scenario = sheet.cell(row=i, column=1)
            if row_scenario.value == scenario_name:
                for j in range(1, sheet.max_column + 1):
                    column_header = sheet.cell(row=1, column=j)
                    if column_header.value == header_name:
                        data = sheet.cell(row=i, column=j)
                        return data.value

    @keyword
    # retrieve all the values from a column and return as a list
    def read_all_column_values(self, sheet_name, column_name):
        df = pd.read_excel("TestData/CsvFiles/Test_Data.xlsx", sheet_name=sheet_name)
        mylist = df[column_name].tolist()
        return mylist

    @keyword
    def payload_to_dict(self, scenario_name):
        data = CustomKeyword.read_data_excel(self, sheet_name="API", scenario_name=scenario_name, header_name="Payload")
        return json.loads(data)

    @keyword
    def payload_to_dict2(self, scenario_name):
        data = scenario_name
        return json.loads(data)

    @keyword
    def string_to_payload(self, scenario_name):
        data = ast.literal_eval(scenario_name)
        return json.loads(data)

    @keyword
    def user_id_from_response(self, response):
        dict2 = response.get('message')
        dict3 = json.loads(dict2)
        dict4 = dict3['result']
        CustomKeyword.write_data_excel(self, sheet_name="API", scenario_name="Login", header_name="UserID",
                                       value_to_write=dict4['id'])
        return dict4['id']

    @keyword
    def token_from_response(self, response):
        dict2 = response.get('message')
        dict3 = json.loads(dict2)
        dict4 = dict3['result']
        CustomKeyword.write_data_excel(self, sheet_name="API", scenario_name="Login", header_name="Auth_Token",
                                       value_to_write=dict4['token'])
        return dict4['token']

    @keyword
    def channelID_from_response(self, response):
        dict2 = response.get('group')
        CustomKeyword.write_data_excel(self, sheet_name="API", scenario_name="DeleteChannel", header_name="Payload",
                                       value_to_write=dict2.get('_id'))
        return dict2.get('_id')

    @keyword
    def convert_to_lower_case(self, stringlower):
        return stringlower.lower()

    @keyword
    def get_os(self):
        os_name = sys.platform
        if "win32" == os_name or "win64" == os_name:
            return "win"
        else:
            return "mac"

    @keyword
    def call_paylaod_verifier(self, response_call):
        #  response_call = json.loads(response_call)
        dict = response_call.get('data')
        providerName = dict.get('providerName')
        if providerName == "jitsi":
            return True
        else:
            return False

    @keyword
    def messageId_from_payload(self, response):
        dict = response.get("message")
        return dict.get('_id')

    @keyword
    def chat_emoji_success(self, response):
        dict = response.get("success")
        return dict

    @keyword
    def text_to_speech(self, text_to_say):
        engine = pyttsx3.init()
        engine.say(text_to_say)
        engine.runAndWait()

    @keyword
    def mpTHREE_to_text(self, audio_path):
        # convert mp3 file to wav
        sound = AudioSegment.from_mp3(audio_path)
        AUDIO_FILE = "TestDownloads/transcript.wav"
        sound.export(AUDIO_FILE, format="wav")

        # use the audio file as the audio source
        r = sr.Recognizer()
        try:
            with sr.AudioFile(AUDIO_FILE) as source:
                audio = r.record(source)  # read the entire audio file
                r.recognize_google(audio, language='en-IN', show_all=True)
                return r.recognize_google(audio)
        except Exception as e:
            return "Exception: " + str(e)

    @keyword
    def enter_button(self):
        # windows keyboard enter key using pyautogui
        pyautogui.press('enter')  # Presses enter

# x = CustomKeyword()
# # x.text_to_speech("welcome to text to speech trial")
# print(x.mpTHREE_to_text("/Users/amananand7/Documents/GitHub/POD5_Mobile_libs/TestDownloads/IlwVCwJ4Y0.mp3"))
